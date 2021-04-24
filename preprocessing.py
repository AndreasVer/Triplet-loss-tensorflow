import os
from matplotlib.image import imread
import numpy as np
import copy
import random
import cv2
import openpyxl


class PreProcessing:
    images_train = np.array([])
    images_test = np.array([])
    labels_train = np.array([])
    labels_test = np.array([])
    unique_train_label = np.array([])
    map_train_label_indices = dict()

    def __init__(self, data_src):

        self.data_src = data_src
        myworkbook = openpyxl.load_workbook("TripletLoss_Sink_2.xlsx")
        worksheet = myworkbook.get_sheet_by_name('Blad1')

        # wb = xlrd.open_workbook("TripletLoss_Sink_2.xlsx")
        # wb= openpyxl.load_workbook(path)
        # sheet = wb.sheet_by_index(0)

        self.anchors = []
        self.positives = {}
        self.negatives = {}
        self.images = {}
        self.placeholder_shape = [64, 64, 3]

        for rownr in range(1,278):
            print(rownr)
            imagename = str(int(worksheet.cell(row=rownr, column=1).value))

            if len(imagename) == 5:
                imagename = "0" + imagename
            self.anchors.append(str(imagename))

            self.images[imagename] = self.normalize(cv2.resize(cv2.imread("sink/" + imagename + ".jpg"), (64, 64), cv2.INTER_AREA))

        for rownr in range(1,278):
            all_imgs = copy.deepcopy(self.anchors)

            imagename = str(int(worksheet.cell(row=rownr, column=1).value))
            if len(imagename) == 5:
                imagename = "0" + imagename

            pos = str(worksheet.cell(row=rownr, column=2).value).replace(" ", "").split(",")
            for p in range(len(pos)):
                if len(pos[p]) == 5:
                    pos[p] = "0" + pos[p]
            self.positives[imagename] = pos
            print(pos, "pos")
            for p in pos:

                print('p', p)
                all_imgs.remove(p)


            self.negatives[imagename] = all_imgs

        print('Preprocessing Done.')

    def normalize(self, x):
        min_val = np.min(x)
        max_val = np.max(x)
        x = (x - min_val) / (max_val - min_val)
        return x

    def read_dataset(self):
        X = []
        y = []
        for directory in os.listdir(self.data_src):
            try:
                for pic in os.listdir(os.path.join(self.data_src, directory)):
                    img = imread(os.path.join(self.data_src, directory, pic))

                    X.append(np.squeeze(np.asarray(img)))
                    y.append(directory)
            except Exception as e:
                print('Failed to read images from Directory: ', directory)
                print('Exception Message: ', e)
        print('Dataset loaded successfully.')
        return X, y

    def preprocessing(self, train_test_ratio):
        X, y = self.read_dataset()
        labels = list(set(y))
        label_dict = dict(zip(labels, range(len(labels))))
        Y = np.asarray([label_dict[label] for label in y])
        X = [self.normalize(x) for x in X]

        shuffle_indices = np.random.permutation(np.arange(len(y)))
        x_shuffled = []
        y_shuffled = []
        for index in shuffle_indices:
            x_shuffled.append(X[index])
            y_shuffled.append(Y[index])

        size_of_dataset = len(x_shuffled)
        n_train = int(np.ceil(size_of_dataset * train_test_ratio))
        return np.asarray(x_shuffled[0:n_train]), np.asarray(x_shuffled[n_train + 1:size_of_dataset]), np.asarray(
            y_shuffled[0:n_train]), np.asarray(y_shuffled[
                                               n_train + 1:size_of_dataset])

    def get_triplets(self):
        a = random.choice(self.anchors)
        p = random.choice(self.positives[a])
        n = random.choice(self.negatives[a])

        return self.images[a], self.images[p], self.images[n]

    def get_triplets_batch(self, n):
        idxs_a, idxs_p, idxs_n = [], [], []
        for _ in range(n):
            a, p, n = self.get_triplets()
            idxs_a.append(a)
            idxs_p.append(p)
            idxs_n.append(n)
        return idxs_a, idxs_p, idxs_n